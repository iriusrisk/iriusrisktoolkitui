# -*- coding: utf-8 -*-
import sys
import os
import pandas as pd
from pathlib import Path
import datetime
import yaml

import textile
import html
import logging
from lxml import etree as etree_

import openpyxl
from src.convertXmlToExcel import applyFormatLibraryRules, applyFormatLibraryProperties, applyFormatLibrary

import src.sample_lib as supermod
from src.xmlValidator import xmlValidationCheck
from src.securityContent import SecurityContent
from src.libraryDetails import readInfoFromXml
from src.mitigationValidator import libMitigationTest
from src.common import exportLib2XML

#Create and configure logger 
#filemode = 'w' that will change the mode of operation from "append" to "write" and will overwrite the file every time we run our application
logging.basicConfig(filename="logFile.log", 
                    format= '%(asctime)s  %(levelname)-10s %(message)s', 
                    datefmt =  "%Y-%m-%d-%H-%M-%S", 
                    filemode='w') 
  
#Creating an object 
logger=logging.getLogger() 
  
#Setting the threshold of logger to INFO 
logger.setLevel(logging.INFO) 

def createProject(scObject, productProperties, componentDefinitions, supportedStandards, rules):
    libName = productProperties['Library Name']['Values']
    libRef = productProperties['Library Ref']['Values']
    libDesc = productProperties['Library Desc']['Values']

    logger.info("Creating library with Name: " + libName + " and Reference: " + libRef)
    rootClass = supermod.libraryType()
    udtsClass = supermod.udtsType()
    dataflowsClass = supermod.dataflowsType()
    componentDefinitionsClass = supermod.componentDefinitionsType()
    supportedStandardsClass = supermod.supportedStandardsType()
    componentsLibraryClass = supermod.componentsLibraryType()
    rulesClass = supermod.rulesType()

    categoryComponentsClass = createCategoryComponents(componentDefinitions)
    componentDefinitionsClass = createComponentDefinitions(componentDefinitions)
    supportedStandardsClass = createSupportedStandards(supportedStandards)

    componentsLibraryClass = createComponents(scObject)
    rulesClass = createRules(rules)

    rootClass.set_udts(udtsClass)
    rootClass.set_dataflows(dataflowsClass)
    rootClass.set_categoryComponents(categoryComponentsClass)
    rootClass.set_supportedStandards(supportedStandardsClass)
    rootClass.set_components(componentsLibraryClass)
    rootClass.set_rules(rulesClass)

    rootObj = rootClass.factory(
        name=libName, 
        ref=libRef, 
        revision=1, type_='LIBRARY', status='OPEN', 
        enabled='true', priority=0, tags='', workflowState='', 
        desc=libDesc, 
        udts=udtsClass, 
        dataflows=dataflowsClass,
        categoryComponents=categoryComponentsClass,
        componentDefinitions=componentDefinitionsClass,
        supportedStandards=supportedStandardsClass,
        components=componentsLibraryClass,
        rules=rulesClass)

    return rootObj

def createConditions(ruleClass, conditions):
    conditionClass=supermod.conditionType()
    patternClass=supermod.patternType()
    names=conditions['Condition Name'].values
    patternNames=conditions['Condition Pattern Name'].values
    patterns=conditions['Condition Pattern Value'].values
    values = conditions['Condition Value'].values
    types=conditions['Condition Type'].values
    fields=conditions['Condition Field'].values
    
    for name, patternName, pattern, value, itype, field in zip(names, patternNames, patterns, values, types, fields):
        if name != '':
            patternObj=patternClass.factory(name=patternName, pattern=pattern)
            conditionObj=conditionClass.factory(name=name, pattern=patternObj, value=value, type=itype, field=field)
            conditionObj.set_type(itype)
            ruleClass.add_condition(conditionObj)
    logger.info("Rule conditions added to the rule '%s'" %ruleClass.get_name())
    return ruleClass

def createActions(ruleClass, actions):
    actionClass=supermod.actionType()
    patternClass=supermod.patternType()
    names=actions['Action Name'].values
    patternNames=actions['Action Pattern Name'].values
    patterns=actions['Action Pattern Value'].values
    values = actions['Action Value'].values
    types=actions['Action Type'].values
    projects=actions['Action Project'].values
    for name, patternName, pattern, value, itype, project in zip(names, patternNames, patterns, values, types, projects):
        if name != '':
            patternObj=patternClass.factory(name=patternName, pattern=pattern)
            actionObj=actionClass.factory(name=name, pattern=patternObj, value=value, type=itype, project=project)
            actionObj.set_type(itype)
            ruleClass.add_action(actionObj)
    logger.info("Rule actions added to the rule '%s'" %ruleClass.get_name())
    return ruleClass

def createRules(rules):
    logger.info("Rule creation started...")
    rulesClass=supermod.rulesType()
    ruleClass=supermod.ruleType()
    if len(rules.values) > 0:
        names = rules['Rule Name'].drop_duplicates().values.tolist()
        names = [name for name in names if str(name) != 'nan']
        modules = rules['Module'].values.tolist()
        modules = [module for module in modules if str(module) != 'nan']
        generatedByGuis = rules['Generated by GUI'].values.tolist()
        generatedByGuis = [generatedByGui for generatedByGui in generatedByGuis if str(generatedByGui) != 'nan']
        # The condition to fill the data with N/A values was wrong (with value: backfill) .The correct one is with ffill.
        index=rules['Rule Name'].fillna(method="ffill")
        rules=rules.fillna("")
        
        conditionColumns=["Condition Name", "Condition Pattern Name", "Condition Pattern Value", "Condition Type", "Condition Value", "Condition Field"]
        actionColumns=["Action Name", "Action Pattern Name", "Action Pattern Value", "Action Type", "Action Value", "Action Project"]
        total=conditionColumns
        total.extend(actionColumns)
        data=rules[total]
        data=pd.concat([index, data], axis=1)
        for name, module, generatedByGui in zip(names, modules, generatedByGuis):
            
            ruleObj = ruleClass.factory(name=name, module=module, generatedByGui=generatedByGui, content="")
            createConditions(ruleObj, data.loc[data['Rule Name']==name])
            createActions(ruleObj, data.loc[data['Rule Name']==name])
            rulesClass.add_rule(ruleObj)
            logger.info("Rule created: '%s'" %name)
    else:
        logger.info("No rules found!")
    return rulesClass


def createSupportedStandards(supportedStandards):
    supportedStandardsClass = supermod.supportedStandardsType()
    supportedStandardClass = supermod.supportedStandardType()

    names = supportedStandards['Supported Standard Name'].drop_duplicates().values.tolist()
    refs = supportedStandards['Supported Standard Ref'].drop_duplicates().values.tolist()
    

    for name, ref in zip(names, refs):
        supportedStandardObj = supportedStandardClass.factory(name=name, ref=ref)
        supportedStandardsClass.add_supportedStandard(supportedStandardObj)
    
    return supportedStandardsClass

def createCategoryComponents(componentDefinitions):
    categoryComponentsClass = supermod.categoryComponentsType()
    categoryComponentClass = supermod.categoryComponentType()

    names = componentDefinitions['Category Name'].drop_duplicates().values.tolist()
    refs = componentDefinitions['Category Ref'].drop_duplicates().values.tolist()
    

    for name, ref in zip(names, refs):
        categoryComponentObj = categoryComponentClass.factory(name=name, ref=ref)
        categoryComponentsClass.add_categoryComponent(categoryComponentObj)

    return categoryComponentsClass

def createComponentDefinitions(componentDefinitions):
    componentDefinitionsClass = supermod.componentDefinitionsType()
    componentDefinitionClass = supermod.componentDefinitionType()   

    names = componentDefinitions['Component Definition Name'].values.tolist()
    refs = componentDefinitions['Component Definition Ref'].values.tolist()
    descs = componentDefinitions['Component Definition Desc'].values.tolist()
    categoryRefs = componentDefinitions['Category Name'].values.tolist()
    categoryRefs = componentDefinitions['Category Ref'].values.tolist()
    riskPatterns = componentDefinitions['Risk Patterns'].values.tolist()

    for name, ref, desc, categoryRef, riskPattern in zip(names, refs, descs, categoryRefs, riskPatterns):
        #Each componentDefinition can import several riskPatterns
        riskPatternsList=list()
        while riskPattern.find(",") != -1:
            riskPatternsList.append(riskPattern[0:riskPattern.find(",")].strip())
            riskPattern=riskPattern[riskPattern.find(",")+1:]
        riskPatternsList.append(riskPattern.strip())
        riskPatternsClass = createRiskPatternsComponentDefinition(riskPatternsList)
        componentDefinitionObj = componentDefinitionClass.factory(ref, name, desc, categoryRef, riskPatternsClass)
        componentDefinitionsClass.add_componentDefinition(componentDefinitionObj)

    return componentDefinitionsClass

def createRiskPatternsComponentDefinition(refs):
    riskPatternsClass = supermod.riskPatternsType()
    riskPatternClass = supermod.riskPatternType()


    for ref in refs:
        riskPatternObj = riskPatternClass.factory(ref)
        riskPatternsClass.add_riskPattern(riskPatternObj)

    return riskPatternsClass

def createComponents(scObject):
    componentsLibraryClass = supermod.componentsLibraryType()
    componentLibraryClass = supermod.componentLibraryType()

    refs = scObject.getComponentIds()
    names = scObject.getComponentNames()

    logger.info("Component creation process STARTED.")
    for i, (name, ref) in enumerate(zip(names, refs)):
        logger.info("Creating component with Name: " + str(name) + " and Reference: " + str(ref))
        #We create weaknesses for each component Ref
        weaknessRefs = scObject.getWeaknessIdsFromComponentId(ref)
        weaknessNames = scObject.getWeaknessNamesFromComponentId(ref)
        weaknessDescs = scObject.getWeaknessDescsFromComponentId(ref)
        weaknessesLibraryClass = createWeaknesses(weaknessRefs, weaknessNames, weaknessDescs)
        componentLibraryClass.set_weaknesses(weaknessesLibraryClass)
        #We create controls for each component Ref
        controlRefs = scObject.getControlIdsFromComponentId(ref)
        controlNames = scObject.getControlNamesFromComponentId(ref)
        controlDescs = scObject.getControlDescsFromComponentId(ref)
        controlTestsSteps = scObject.getControlTestsStepsFromComponentId(ref)
        referencesForControls = scObject.getControlReferencesFromComponentId(ref)
        standardsForControls = scObject.getControlStandardsFromComponentId(ref)
        controlsLibraryClass = createControls(controlRefs, controlNames, controlDescs, controlTestsSteps, referencesForControls, standardsForControls)
        componentLibraryClass.set_controls(controlsLibraryClass)
        #We create use cases for each component Ref. Usecases hold the relationship between threats, weaknesses and controls.
        useCasesNames = scObject.getUseCaseNamesFromComponentId(ref)
        usecasesLibraryClass = createUseCases(useCasesNames, ref, scObject)
        componentLibraryClass.set_usecases(usecasesLibraryClass)

        componentLibraryObj = componentLibraryClass.factory(ref, name, desc='', 
            groupName='', tags='', position=i+1, library='', 
            diagramPositionX='0', diagramPositionY='0', 
            componentDefinitionRef='',
            weaknesses=weaknessesLibraryClass,
            controls=controlsLibraryClass,
            usecases=usecasesLibraryClass)
        componentsLibraryClass.add_component(componentLibraryObj)

    logger.info("Component creation process ENDED.")

    return componentsLibraryClass

def createWeaknesses(refs, names, descs):
    weaknessesLibraryClass = supermod.weaknessesLibraryType()
    weaknessLibraryClass = supermod.weaknessLibraryType()

    for name, ref, desc in zip(names, refs, descs):
        if ref != "":
            logger.info("Creating weakness with Name: " + name + " and Reference: " + ref)
            testLibraryObj = createTest()
            weaknessLibraryObj = weaknessLibraryClass.factory(ref=ref, name=name, state=0, desc=desc, impact=100, test=testLibraryObj)
            weaknessesLibraryClass.add_weakness(weaknessLibraryObj)

    return weaknessesLibraryClass

def createControls(refs, names, descs, testsSteps, referencesForControls, standardsForControl):
    controlsLibraryClass = supermod.controlsLibraryType()
    controlLibraryClass = supermod.controlLibraryType()
    implementationsClass = supermod.implementationsType()
    standardsClass = supermod.standardsType()
    udtsClass = supermod.udtsType()

    for name, ref, desc, testSteps, referencesForControl, standardsForControl in zip(names, refs, descs, testsSteps, referencesForControls, standardsForControl):
        logger.info("Creating control with Name: " + name + " and Reference: " + ref + " and URL References: " + referencesForControl)
        referencesClass = createReferences(referencesForControl)
        standardsClass = createStandards(standardsForControl)
        testLibraryObj = createTest(testSteps=testSteps)
        controlLibraryObj = controlLibraryClass.factory(ref=ref, name=name, desc=convertText2Html(desc), platform='', 
            cost='1', risk='0', state='Recommended', owner='', library='', source='MANUAL', test=testLibraryObj, 
            implementations= implementationsClass, references=referencesClass, standards=standardsClass, udts=udtsClass)
        controlsLibraryClass.add_control(controlLibraryObj)

    return controlsLibraryClass

def createUseCases(names, componentId, scObject):
    usecasesLibraryClass = supermod.usecasesLibraryType()
    usecaseLibraryClass = supermod.usecaseLibraryType()

    for name in names:
        name = str(name)
        ref = name.upper()
        logger.info("Creating Use Case with Reference: " + ref)
        #We create threats for each useCase Ref
        threatRefs = scObject.getThreatIdsFromComponentIdAndUseCase(componentId, name)
        threatNames = scObject.getThreatNamesFromComponentIdAndUseCase(componentId, name)
        threatDescs = scObject.getThreatDescsFromComponentIdAndUseCase(componentId, name)
        threatsLibraryClass = createThreats(componentId, threatRefs, threatNames, threatDescs, scObject)
        usecaseObj = usecaseLibraryClass.factory(name=name, ref=ref, desc='', library='', threats=threatsLibraryClass)
        usecasesLibraryClass.add_usecase(usecaseObj)

    return usecasesLibraryClass

def createThreats(componentId, refs, names, descs, scObject):
    threatsLibraryClass = supermod.threatsLibraryType()
    threatLibraryClass = supermod.threatLibraryType()
    riskRatingClass = supermod.riskRatingType()
    referencesClass = supermod.referencesType()

    riskRatingObj = riskRatingClass.factory(confidentiality='100', integrity='100', availability='100', easeOfExploitation='25')

    for name, ref, desc in zip(names, refs, descs):
        #Each Threat have its related controls (controlsRef)
        controlRefs = scObject.getControlIdsFromThreatIdAndComponentId(ref, componentId)
        #controlsDict dictionary stores (key,value) pairs of (controlRef, mitigation) values
        controlsDict = scObject.calculateControlsMitigations(controlRefs)
        controlsRefClass = createControlsRef(controlsDict)
        #showControlsRef(controlsRefClass)
        #We duplicate the relation between Threat -> Control with Threat -> Weakness -> Control
        weaknessRefs = scObject.getWeaknessIdsFromThreatIdAndComponentId(ref, componentId)
        weaknessesRefClass = createWeaknessesRef(componentId, ref, weaknessRefs, scObject, controlsDict)

        threatLibraryObj = threatLibraryClass.factory(ref=ref, name=name, desc=convertText2Html(desc), state='Expose', source='MANUAL', owner='', library='',
            riskRating=riskRatingObj, references=referencesClass, controls=controlsRefClass, weaknesses=weaknessesRefClass)
        threatsLibraryClass.add_threat(threatLibraryObj)

    return threatsLibraryClass


def createControlsRef(controlsDict):
    controlsRefClass = supermod.controlsRefType()
    controlRefClass = supermod.controlRefType()

    for ref, mitigation in controlsDict.items():
        controlRefObj = controlRefClass.factory(ref=ref, mitigation=mitigation)
        controlsRefClass.add_control(controlRefObj)

    return controlsRefClass

def showControlsRef(controlsRefClass):
    
    for control in controlsRefClass.get_control():
        ref = control.get_ref()
        mitigation = control.get_mitigation()
        print("----> Control Reference: " + ref)
        print("----> Control Mitigation: " + str(mitigation))


def createWeaknessesRef(componentId, threatId, refs, scObject, controlsDict):
    weaknessesRefClass = supermod.weaknessesRefType()
    weaknessRefClass = supermod.weaknessRefType()

    for ref in refs:
        if ref != "":
            controlRefs = scObject.getControlIdsFromWeaknessIdAndThreatIdAndComponentId(ref, threatId, componentId)
            filteredControlsDict = scObject.getFilteredControlsDictByControlRefs(controlsDict, controlRefs)
            controlsRefClass = createControlsRef(filteredControlsDict)
            weaknessRefObj = weaknessRefClass.factory(ref=ref, controls=controlsRefClass)
            weaknessesRefClass.add_weakness(weaknessRefObj)

    return weaknessesRefClass

def createStandards(standardsForControl):
    standardsClass = supermod.standardsType()
    standardClass = supermod.standardType()

    #referencesForControl is not a Python list, it's a string that need to be parsed to find the actual references
    if(standardsForControl.find("[")!=-1):
        while(standardsForControl.find("[")!=-1):
            if(standardsForControl.find("]")==-1 or standardsForControl.find("|")==-1):
                print("Error creating Control standard from: " + standardsForControl)
                break
            supportedStandardRef=standardsForControl[standardsForControl.find("[")+1:standardsForControl.find("|")]
            ref=standardsForControl[standardsForControl.find("|")+1:standardsForControl.find("]")]
            standardObj = standardClass.factory(supportedStandardRef=supportedStandardRef, ref=ref)
            standardsClass.add_standard(standardObj)
            standardsForControl=standardsForControl[standardsForControl.find("]")+1:]
    else:
        logger.warning("No standards to add for Control")

    return standardsClass

def createReferences(referencesForControl):
    referencesClass = supermod.referencesType()
    referenceClass = supermod.referenceType()

    #referencesForControl is not a Python list, it's a string that need to be parsed to find the actual references
    if(referencesForControl.find("[")!=-1):
        while(referencesForControl.find("[")!=-1):
            if(referencesForControl.find("]")==-1 or referencesForControl.find("|")==-1):
                print("Error creating Control URL References from: " + referencesForControl)
                break
            ref_name=referencesForControl[referencesForControl.find("[")+1:referencesForControl.find("|")]
            ref_url=referencesForControl[referencesForControl.find("|")+1:referencesForControl.find("]")]
            referenceObj = referenceClass.factory(ref_name, ref_url)
            referencesClass.add_reference(referenceObj)
            referencesForControl=referencesForControl[referencesForControl.find("]")+1:]
    else:
        logger.warning("No references to add for Control")

    return referencesClass

def createTest(testSteps=''):
    testLibraryClass = supermod.testLibraryType()
    referencesClass =supermod.referencesType()
    sourceClass = supermod.sourceType()
    timestamp = datetime.datetime.now().replace(microsecond=0).isoformat()
    sourceObj = sourceClass.factory(filename='', args='', type_='Manual', result='Not Tested', enabled=True, timestamp=timestamp, output='')
    udtsClass = supermod.udtsType()
    testLibraryObj = testLibraryClass.factory(expiryDate='', expiryPeriod=0, steps=convertText2Html(testSteps), notes='', 
        references=referencesClass, source=sourceObj, udts=udtsClass)
    return testLibraryObj
    


def convertText2Html(text):
    
    text = textile.textile(text)
    
    text=text.replace("<span class=\"caps\">","")
    text=text.replace("</span>","")

    return text 


USAGE_TEXT = """
You should pass more args!!
Please give the arguments in the following order:
1 - The file name of the Excel file with the extension (.xls , .xlsx, ...).
2 - The sheet name ofthe Excel file.
Usage: python sample_app1.py test.xlsx SheetName
"""

def usage():
    print(USAGE_TEXT)
    sys.exit(1)

def generateXmlFromXlsxFile(excelFilePath, productProperties, componentDefinitions, supportedStandards, rules, error):
    sc = SecurityContent(logger)
    sc.importExcel(excelFilePath)
    
    xmlFile = str(productProperties['Library Ref']['Values']).replace(" ","-") + '.xml'
    xmlFileName = Path.cwd() / "outFiles" / "outputLibs" / xmlFile
    xsdFileName = Path.cwd() / "inputFiles" / "XSD_Schema" / "library.xsd"
    
    rootObj = createProject(sc, productProperties, componentDefinitions, supportedStandards, rules)

    exportLib2XML(xmlFileName, rootObj)

    if(productProperties['Do Lib validations']['Values'] =='Yes'):
        print("-> Lib validations: ON")
        print("-- (1/2) Schema validation --")
        xmlValidationCheck(str(xmlFileName), str(xsdFileName))
        print("-- (2/2) Mitigation validation --")
        libMitigationTest(str(xmlFileName), [])
    else:
        print("-> Lib validations: OFF")

    if(productProperties['Show stats']['Values'] =='Yes'):
        print("-> Library Statistics: ON")
        columns=['Library Name', 'Risk Pattern', '# Use Cases', '# Threats', '# Weaknesses', '# Countermeasures']
        readInfoFromXml(xmlFileName, columns)
    else:
        print("-> Library Statistics: OFF")
    
    return xmlFileName, error

def generateXmlFromRulesXlsxFile(excelFilePath, productProperties, componentDefinitions, supportedStandards, rules):
    sc = SecurityContent(logger)
    sc.setEmptySecurityContent()
    
    xmlFile = str(productProperties['Library Ref']['Values']).replace(" ","-") + '.xml'
    xmlFileName = Path.cwd() / "outFiles" / "outputLibs" / xmlFile
    xsdFileName = Path.cwd() / "inputFiles" / "XSD_Schema" / "library.xsd"
    
    rootObj = createProject(sc, productProperties, componentDefinitions, supportedStandards, rules)

    exportLib2XML(xmlFileName, rootObj)

    if(productProperties['Do Lib validations']['Values'] =='Yes'):
        print("-> Lib validations: ON")
        print("-- Schema validation --")
        xmlValidationCheck(str(xmlFileName), str(xsdFileName))
    else:
        print("-> Lib validations: OFF")
    
    return xmlFileName



def checkIfRulesTabExists(excelPath):
    errors=""
    excel = openpyxl.load_workbook(str(excelPath))
    if not "Rules" in excel.get_sheet_names():
        logger.info("The tab Rules doesn't exist in the Excel with path: %s"%excelPath)
        excel.create_sheet("Rules")
        columns=['Rule Name', 'Module',	'Generated by GUI',	'Condition Name', 'Condition Pattern Name',	'Condition Pattern Value', 'Condition Type', 'Condition Value', 'Condition Field', 'Action Name', 'Action Pattern Name', 'Action Pattern Value', 'Action Type', 'Action Value', 'Action Project']
        sheet = excel.get_sheet_by_name('Rules')
        for col in columns:
            sheet.cell(row = 1, column = columns.index(col)+1).value = str(col)        
        excel.save(str(excelPath))

        applyFormatLibraryRules(str(excelPath))
        errors="The Rules tab was created, because it wasn't found.\n"
    return errors

def checkTabsInExcelRulesEditor(excelPath):
    #We check the Excel Workbook have the ringt WorkSheets inside
    errors=""

    xl = pd.ExcelFile(str(excelPath))

    for tab in ["Rules", "Conditions", "Actions", "Lists", "RiskPatterns", "Library properties"]:
        if not tab in xl.sheet_names:
            logger.info("The tab Rules doesn't exist in the Excel with path: %s" % excelPath)
            errors="Worksheet: %s was not found in the Excel Workbook: %s.\n" % (tab, excelPath)
    return errors


def checkIfAllColumnsExists(excelPath):
    # We check if the number of columns in the Risk Pattern tab is corretct
    columnsFila1=['Risk Pattern', '', '','Use case', 'Threat', '', '', '', 'Weakness', '', '', '', 'Countermeasure', '', '', '', '', '']
    columnsFila2=['Id', 'Name', 'Desc', 'Name', 'Id', 'Name', 'Desc', 'References', 'Id', 'Name', 'Desc', 'References', 'Id', 'Name', 'Desc', 'Test steps', 'References', 'Standards']
    dfm=pd.read_excel(str(excelPath), sheet_name="Risk Patterns", header=None)
    dfm=dfm.fillna("")
    columns1=dfm.iloc[0].replace("nan","")
    columns2=dfm.iloc[1].replace("nan","")
    cont=0
    addColumns=list()
    cols=list()
    for i in range(0,len(columnsFila1)):
        iter1=str("%s %s"%(columns1[cont], columns2[cont])).strip()
        iter2=str("%s %s"%(columnsFila1[i], columnsFila2[i])).strip()        
        if iter1 == iter2:
            cols.append(iter1)
            cont+=1
        else:
            addColumns.append(i)
    
    # If some column is needed to be added, we shall get the data from the Rules and Properties tabs to regenerate the Excel file.
    if len(addColumns) != 0:
        long= len(dfm.iloc[:,0])
        column = ["" for i in range(0,long)]
        dfm.columns=cols
        error="Please, fill the empty columns in the Excel file with path: '%s'. And run the conversion of Excel to XML again.\n"%excelPath
        for it in addColumns:
            column[0]=columnsFila1[it]
            column[1]=columnsFila2[it]
            dfm.insert(it, str("%s %s"%(columnsFila1[it], columnsFila2[it])).strip(), column, True)  
        dfmRules=pd.read_excel(str(excelPath), sheet_name='Rules', header=None)
        dfmProperties=pd.read_excel(str(excelPath), sheet_name='Library properties', header=None)
        with pd.ExcelWriter(str(excelPath)) as writer:
            dfm.to_excel(excel_writer=writer, sheet_name="Risk Patterns", header=None, index=None)
            dfmRules.to_excel(excel_writer=writer, sheet_name="Rules", header=None, index=None) 
            dfmProperties.to_excel(excel_writer=writer, sheet_name="Library properties", header=None, index=None) 
        
        applyFormatLibrary(str(excelPath), sheetName='Risk Patterns')
        applyFormatLibraryProperties(str(excelPath))
        applyFormatLibraryRules(str(excelPath))

    else:
        error=""
    
    return error
  

def getRulesFromExcel(excelPath):
    errors=checkIfRulesTabExists(excelPath)
    return pd.read_excel(str(excelPath), sheet_name="Rules", columns=0), errors

def getRulesFromExcelRulesEditor(excelPath):
    errors = checkTabsInExcelRulesEditor(excelPath)
    all_dfs = pd.read_excel(excelPath, sheet_name=None)

    #We create an empty datarame for the rules
    columns = ['Rule Name', 'Module', 'Generated by GUI', 
    'Condition Name', 'Condition Pattern Name', 'Condition Pattern Value', 'Condition Type', 'Condition Value', 'Condition Field', 
    'Action Name', 'Action Pattern Name', 'Action Pattern Value', 'Action Type', 'Action Value', 'Action Project']
    columns_1 = ['Rule Name', 'Module', 'Generated by GUI']
    columns_2 = ['Rule Name', 'Condition Name', 'Condition Pattern Name', 'Condition Pattern Value', 'Condition Type', 'Condition Value', 'Condition Field']
    columns_3 = ['Rule Name', 'Action Name', 'Action Pattern Name', 'Action Pattern Value', 'Action Type', 'Action Value', 'Action Project']

    d_1 = dict.fromkeys(columns_1)
    d_2 = dict.fromkeys(columns_2)
    d_3 = dict.fromkeys(columns_3)

    class ConditionPattern:
        QESTION_GROUP_EXISTS = "$group : QuestionGroup(id == \"${value}\") @Watch(!*);"
        QUESTION_IS_ANSWERED = "${type}(${field} == \"${value}\", answer == true);"
        COMPONENT_DEFINITION = "$group : QuestionGroup(id == \"${value}\") @Watch(!*);"

    class ActionPattern:
        INSERT_QUESTION = "insertLogical(new Question($group, \"${value-1}\", \"${value-2}\", \"${value-3}\"));"
        INSERT_QUESTION_GROUP = "insertLogical(new QuestionGroup(\"${value-1}\",\"${value-2}\";, \"${value-3}\", ${value-4}, ${value-5}, ${value-6}, \"${value-7}\"));"
        IMPORT_RISK_PATTERN = "insertLogical(ImportRiskPatternFactory.importRiskPattern(\"${value-1}\",\"${value-2}\", 50));"

    #We will populate rules dataframe using a list of dictionaries.
    rules = []
    conditions = []
    actions = []
    data = []

    for rindex, rrow in all_dfs["Rules"].iterrows():
        #We reset dictionary values
        d_1 = d_1.fromkeys(d_1, None)
        logger.info("Processing Rule Name: %s" % rrow["Rule name"])
        d_1["Rule Name"] = rrow["Rule name"]
        d_1["Module"] = rrow["Rule module type"]
        d_1["Generated by GUI"] = 1
        rules.append(d_1)

    for cindex, crow in all_dfs["Conditions"].iterrows():
        d_2 = d_2.fromkeys(d_2, None)
        d_2["Rule Name"] = crow["Rule name"]
        if crow["Condition"] == "Question Group exists":
            d_2["Condition Name"] = crow["Condition"]
            d_2["Condition Pattern Name"] = "Question group exists"
            d_2["Condition Pattern Value"] = ConditionPattern.QESTION_GROUP_EXISTS
            d_2["Condition Type"] = "drools-without-variable-type"
            d_2["Condition Value"] = crow["QGE-Condition-Value"]
            d_2["Condition Field"] = "id"
        elif crow["Condition"] == "Question is answered":
            d_2["Condition Name"] = crow["Condition"]
            d_2["Condition Pattern Name"] = "Condition with answer"
            d_2["Condition Pattern Value"] = ConditionPattern.QUESTION_IS_ANSWERED
            d_2["Condition Type"] = "Question"
            d_2["Condition Value"] = crow["QIA-Condition-Value"]
            d_2["Condition Field"] = "id"
        elif crow["Condition"] == "Component Definition":
            d_2["Condition Name"] = "Is specific component definition"
            d_2["Condition Pattern Name"] = "Is specific component definition"
            d_2["Condition Pattern Value"] = ConditionPattern.COMPONENT_DEFINITION
            d_2["Condition Type"] = "drools-without-variable-type"
            d_2["Condition Value"] = crow["CD-Condition-Value"]
            d_2["Condition Field"] = "id"
        else:
            errors += "Rule Condition: '%s'. not implemented.\n"%crow["Condition"]

        conditions.append(d_2)

    for aindex, arow in all_dfs["Actions"].iterrows():
        d_3 = d_3.fromkeys(d_3, None)
        d_3["Rule Name"] = arow["Rule name"]
        if arow["Action"] == "Insert Question":
            d_3["Action Name"] = arow["Action"]
            d_3["Action Pattern Name"] = "Action of insert question"
            d_3["Action Pattern Value"] = ActionPattern.INSERT_QUESTION
            d_3["Action Type"] = "drools-without-variable-type"
            d_3["Action Value"] = arow["IQ-QuestionID"] + "_::_" + arow["IQ-QuestionOption"] + "_::_" + arow["IQ-Description"]
            d_3["Action Project"] = ""
        elif arow["Action"] == "Insert Question Group":
            d_3["Action Name"] = arow["Action"]
            d_3["Action Pattern Name"] = "Action of insert question group"
            d_3["Action Pattern Value"] = ActionPattern.INSERT_QUESTION_GROUP
            d_3["Action Type"] = "drools-without-variable-type"
            d_3["Action Value"] = arow["IQG-QuestionGroupID"] + "_::_" + arow["IQG-Category"] + "_::_" + arow["IQG-QuestionText"] + "_::_"\
            + str(int(arow["IQG-Priority"])) + "_::_" + str(arow["IQG-MutexAnswersFlag"]).lower() + "_::_" + str(arow["IQG-RequiredFlag"]).lower() + "_::_"\
            + arow["IQG-DescriptionText"]
            d_3["Action Project"] = ""
        elif arow["Action"] == "Import Risk Pattern":
            d_3["Action Name"] = arow["Action"]
            d_3["Action Pattern Name"] = "Action of import a risk pattern"
            d_3["Action Pattern Value"] = ActionPattern.IMPORT_RISK_PATTERN
            d_3["Action Type"] = "drools-without-variable-type"
            d_3["Action Value"] = arow["IRP-LibraryID"] + "_::_" + arow["IRP-RiskPatternID"]
            d_3["Action Project"] = arow["IRP-LibraryID"]
        else:
            errors += "Rule Action: '%s'. not implemented.\n"%arow["Action"]

        actions.append(d_3)

    for rule in rules:
        row = rule
        ruleName = row["Rule Name"]

        conditionsForRule = filterListByRuleName(conditions, ruleName)
        actionsForRule = filterListByRuleName(actions, ruleName)
        number_rows = calculateNumberOfRows(conditionsForRule, actionsForRule)

        for i in range(number_rows):
            if i < len(conditionsForRule):
                row.update(conditionsForRule[i])
            if i < len(actionsForRule):
                row.update(actionsForRule[i])

            data.append(row)
            row = {}

    rules_df = pd.DataFrame(data, columns = columns)
    logger.info("%s rules inseserted" % str(len(rules)))

    return rules_df, errors

def filterListByRuleName(myList, ruleName):
    myListFiltered = []
    for d in myList:
        rName = d.get("Rule Name", "")
        if rName != "":
            if rName == ruleName:
                #We delete key 'rule name' for the dictionary
                d2 = {i:d[i] for i in d if i!="Rule Name"}
                myListFiltered.append(d2)
        else:
            logger.error("Key 'Rule Name' not found in dictionary: %s" % str(d))

    return myListFiltered


def calculateNumberOfRows(conditionsRule, actionsRule):
    number_conditions = len(conditionsRule)
    number_actions = len(actionsRule)
    if number_conditions >= number_actions:
        number_rows = number_conditions
    else:
        number_rows = number_actions

    return number_rows

def getPropertiesFromExcel(excelPath):
    error=checkIfAllColumnsExists(excelPath)
    dfm = pd.read_excel(str(excelPath), sheet_name='Library properties', columns=0)
    
    dataProject = dfm.filter(['General', 'Values']).T
    header = dataProject.iloc[0]
    dataProject = dataProject[1:]
    dataProject=dataProject.rename(columns = header)
    dataProject = dataProject.fillna('')
    
    componentDefinitions=dfm.filter(['Component Definition Name','Component Definition Ref','Component Definition Desc','Category Name', 'Category Ref','Risk Patterns'])
    componentDefinitions=componentDefinitions.dropna(how='all')
    componentDefinitions=componentDefinitions.fillna('')

    supportedStandards=dfm.filter(['Supported Standard Name','Supported Standard Ref'])
    supportedStandards=supportedStandards.dropna(how='all')
    supportedStandards=supportedStandards.fillna('')

    
    return (dataProject, componentDefinitions, supportedStandards, error)

def getSimplifiedPropertiesFromExcel(excelPath):
    
    dfm = pd.read_excel(str(excelPath), sheet_name='Library properties', columns=0)
    
    dataProject = dfm.filter(['General', 'Values']).T
    header = dataProject.iloc[0]
    dataProject = dataProject[1:]
    dataProject=dataProject.rename(columns = header)
    dataProject = dataProject.fillna('')
    #Empty dataframe for componentDefinitions
    componentDefinitions = pd.DataFrame(columns=['Component Definition Name','Component Definition Ref','Component Definition Desc','Category Name', 'Category Ref','Risk Patterns'])
    #Empty dataframe for supportedStandards
    supportedStandards = pd.DataFrame(columns=['Supported Standard Name','Supported Standard Ref'])
    
    return (dataProject, componentDefinitions, supportedStandards)


def convertExcelToXml(excelFilePath):
    rules, errorsRules = getRulesFromExcel(excelFilePath)
    #generateDataFrameHTMLReport(rules)
    (productProperties,componentDefinitions, supportedStandards, errors)=getPropertiesFromExcel(excelFilePath)
    if errors == "":
        xmlFileName, errors = generateXmlFromXlsxFile(excelFilePath, productProperties,componentDefinitions, supportedStandards, rules, errors)
        libraryName = supermod.parse(str(xmlFileName), silence=True).get_name()
        errors+="Library '%s' was converted to XML file successfully and the output file is in the path '%s'.\n"%(libraryName, xmlFileName)
    return errorsRules+errors

def convertRulesFromExcelToXML(excelFilePath):
    rules, errorsRules = getRulesFromExcelRulesEditor(excelFilePath)
    #generateDataFrameHTMLReport(rules)
    (productProperties,componentDefinitions, supportedStandards) = getSimplifiedPropertiesFromExcel(excelFilePath)
    xmlFileName = generateXmlFromRulesXlsxFile(excelFilePath, productProperties,componentDefinitions, supportedStandards, rules)
    libraryName = supermod.parse(str(xmlFileName), silence=True).get_name()
    okMessage = "Rules in Excel Utility '%s' were converted to XML file successfully and the output file is in the path '%s'.\n"%(excelFilePath, xmlFileName)
    logger.info("Excel Rules converted successfully to XML")

    return errorsRules+okMessage

def generateDataFrameHTMLReport(dataframe):
    pd.set_option('display.max_colwidth', -1)
    pd.set_option('max_columns', None)
    output_path = Path.cwd() / "outFiles" / "generatedHtml" / "dataFrameHTMLReport.html"
    dataframe.to_html(str(output_path), escape=False,justify='center', notebook=True, sparsify=False)
    print("You can find a detailed HTML report in the following path:\n" + str(output_path))

# We use this script to catch the arguments and run the principal method
def main():
 
    path=Path.cwd() / "inputFiles" / "spreadSheetFiles"
    excelFiles=os.listdir(str(path))
    text="Select the number of the desired Excel to be converted to xml:\n"
    for excelFile in excelFiles:
        text+="%i - %s\n" %(excelFiles.index(excelFile), excelFile)
    value = 9999
    while value <0 or value>len(excelFiles):
        value=int(input(text))

    excelFilePath = path / excelFiles[value]   
    convertExcelToXml(excelFilePath)
                
        
    

if __name__ == '__main__':
    main()




