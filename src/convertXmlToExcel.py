import os
from pathlib import Path
home=os.getcwd()
import pandas as pd
from lxml import etree

import src.sample_lib as supermod

from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import logging
_log = logging.getLogger(__name__)

# We get the basic data of the library, and we return a list with the data
def getLibraryDetails(rootClass):
  libraryDetails=list()
  libraryDetails.append(["Library Name", rootClass.get_name()])
  libraryDetails.append(["Library Ref", rootClass.get_ref()])
  libraryDetails.append(["Library Desc", rootClass.get_desc()])
  libraryDetails.append(["Do Lib validations", "Yes"])
  libraryDetails.append(["Show stats", "Yes"])
  return libraryDetails

# We recopile all data of the component definitions and we return it as list
def getComponentDefinitionDetails(rootClass):
  componentDefinitionDetails=list()
  categoryComponents=rootClass.get_categoryComponents()
  componentDefinitions=rootClass.get_componentDefinitions()
  for componentDefinition in componentDefinitions.get_componentDefinition():
    componentDefinitionName=componentDefinition.get_name()
    componentDefinitionRef=componentDefinition.get_ref()
    componentDefinitionDesc=componentDefinition.get_desc()
    categoryRef=componentDefinition.get_categoryRef()
    categoryName=""
    for categoryComponent in categoryComponents.get_categoryComponent():
      if categoryComponent.get_ref() == categoryRef:
        categoryName=categoryComponent.get_name()
    riskPatterns=componentDefinition.get_riskPatterns()
    compRiskPatterns=""
    for riskPattern in riskPatterns.get_riskPattern():
      compRiskPatterns+="%s," %riskPattern.get_ref()
    compRiskPatterns=compRiskPatterns[:-1]
    componentDefinitionDetails.append([componentDefinitionName, componentDefinitionRef, componentDefinitionDesc, categoryName, categoryRef, compRiskPatterns])

  return componentDefinitionDetails

# We get all data of the supported standard and we return it as list
def getSupportedStandardDetails(rootClass):
  supportedStandardDetails=list()
  supportedStandards=rootClass.get_supportedStandards().get_supportedStandard()
  for supportedStandard in supportedStandards:
    supportedStandardDetails.append([supportedStandard.get_name(), supportedStandard.get_ref()])
  
  return supportedStandardDetails

def getRulesData(rootClass):
  table=list()
  rules=rootClass.get_rules()
  for rule in rules.get_rule():
    conditions=rule.get_condition()
    actions=rule.get_action()
    ruleDetails=[
      rule.get_name(),
      rule.get_module(),
      rule.get_generatedByGui()
    ]
    minLen=min(len(conditions), len(actions))
    maxLen=max(len(conditions), len(actions))
    for i in range(0, maxLen):
      row=list()      
      for ruleDetail in ruleDetails:
        if i == 0:
          row.append(ruleDetail)
        else:
          row.append("")       
      if i < len(conditions):
        row=addConditionDetailsToRow(conditions[i], row)
      else:
        row=addConditionDetailsToRow("", row)
      if i < len(actions):
        row=addActionDetailsToRow(actions[i], row)
      else:
        row=addActionDetailsToRow("", row)
      table.append(row)

  return table
def addConditionDetailsToRow(condition, row):
  if condition == "":
    row.append("")
    row.append("")
    row.append("")
  else:
    row.append(condition.get_name())
    row.append(condition.get_value())
    row.append(condition.get_field())

  return row

def addActionDetailsToRow(action, row):
  if action == "":
    row.append("")
    row.append("")
    row.append("")
  else:
    row.append(action.get_name())
    row.append(action.get_value())
    row.append(action.get_project())

  return row
# We get all data of the risk pattern from the library, and we return it as list
def getAllDataFromRiskPatterns(rootClass):
  table=list()
  riskPatterns=rootClass.get_components()
  supportedStandards=rootClass.get_supportedStandards()
  for riskPattern in riskPatterns.get_component():
    componentName=riskPattern.get_name()
    componentRef=riskPattern.get_ref()
    componentDesc=convertFormatText(riskPattern.get_desc())
    componentDetail=[componentRef, componentName, componentDesc]
    compControls=riskPattern.get_controls()
    compWeaknesses=riskPattern.get_weaknesses()
    # We get the usecase data
    usecases=riskPattern.get_usecases()

    # Risk patterns that are "not complete" (not having a full path to at least one control)
    # won't appear in the exported Excel. That means if a risk pattern only has the use case / use case+threat / use case+threat+weakness
    # that line won't be written. Take CWE-7-KINGDOMS for example.
    # We could do this to add those risk patterns that doesn't have any use case, but it is not the best solution
    # if len(usecases.get_usecase()) == 0:
    #   table.append(componentDetail + ['-'] * 15)
    # else:

    for usecase in usecases.get_usecase():
      usecaseDetail = [usecase.get_ref(), usecase.get_name(), usecase.get_desc()]
      threats=usecase.get_threats()
      for threat in threats.get_threat():
        # We get the threat data
        threatDetail=[threat.get_ref(), threat.get_name(), convertFormatText(threat.get_desc()), convertRefsToString(threat.get_references())]
        # We get the weakness data
        weaknesses=threat.get_weaknesses() 
        controls=list()
        for weakness in weaknesses.get_weakness():
          weaknessRef=weakness.get_ref()
          for compWeakness in compWeaknesses.get_weakness():
            if compWeakness.get_ref() == weaknessRef:
              weaknessDetail=[weaknessRef, compWeakness.get_name(), convertFormatText(compWeakness.get_desc()), convertRefsToString(compWeakness.get_test().get_references())]
          # We get the countermeasure data
          controls=weakness.get_controls()
          for control in controls.get_control():
            controlRef=control.get_ref()
            for compControl in compControls.get_control():
              if compControl.get_ref() == controlRef:
                controlDetail=[
                  controlRef, 
                  compControl.get_name(), 
                  convertFormatText(compControl.get_desc()), 
                  convertFormatText(compControl.get_test().get_steps()),
                  convertRefsToString(compControl.get_references()), 
                  convertStandardsToString(compControl.get_standards(),supportedStandards)
                  ]
            table.append(componentDetail+usecaseDetail+threatDetail+weaknessDetail+controlDetail)
            componentDetail=["","",""]
            usecaseDetail=["","",""]
            threatDetail=["","","",""]
            weaknessDetail=["","","",""] 
        threatControls=threat.get_controls()
        for control in threatControls.get_control():
          controlRef=control.get_ref()
          if len(weaknesses.get_weakness()) == 0:
            weaknessDetail=["","","",""] 
            for compControl in compControls.get_control():
              if compControl.get_ref() == controlRef:
                controlDetail=[
                  controlRef, 
                  compControl.get_name(), 
                  convertFormatText(compControl.get_desc()), 
                  convertFormatText(compControl.get_test().get_steps()),
                  convertRefsToString(compControl.get_references()), 
                  convertStandardsToString(compControl.get_standards(),supportedStandards)
                  ]
            table.append(componentDetail+usecaseDetail+threatDetail+weaknessDetail+controlDetail)
            componentDetail=["","",""]
            usecaseDetail=["","",""]
            threatDetail=["","","",""]
            weaknessDetail=["","","",""] 
            

            # We reset the data
              
  return table 
# With this method, we want to get all necessary data from the xml file and convert it to pandas DataFrame
def getDataFromXml(xmlPath):
  table=list()
  # Main headers for the Excel
  table.append(["Risk Pattern", "", "", "Use case", "", "", "Threat", "", "", "", "Weakness", "", "", "", "Countermeasure", "", "", "", "", ""])
  table.append(["Id", "Name", "Desc", "Id", "Name", "Desc", "Id", "Name", "Desc", "References", "Id", "Name", "Desc", "References", "Id", "Name", "Desc", "Test steps", "References", "Standards"
  ])
  # We get all nodes of the xml file
  rootClass=supermod.parse(xmlPath, silence=True)
  # we get all important data from the xml file and transform it in a list
  table.extend(getAllDataFromRiskPatterns(rootClass))

  # These are the column names of the DataFrame
  columns=["Risk Pattern Id", "Risk Pattern Name", "Risk Pattern Desc", "Use case Id", "Use case Name", "Use case Desc", "Threat Id", "Threat Name", "Threat Desc", "Threat Refs", "Weakness Id", "Weakness Name", "Weakness Desc", "Weakness Refs", "Countermeasure Id", "Countermeasure Name", "Countermeasure Desc", "Countermeasure Test steps", "Countermeasure Refs", "Countermeasure standards"]
  dfm=pd.DataFrame(table, columns=columns)

  # We get the library details, component definitions and supported standards
  libraryDetails=getLibraryDetails(rootClass)
  componentDefinitionDetails = getComponentDefinitionDetails(rootClass)
  supportedStandardDetails = getSupportedStandardDetails(rootClass)
    
  # We convert all library information to dataframes
  dfmLibrary=pd.DataFrame(libraryDetails, columns=["General", "Values"])
  dfmComponentDefinitions=pd.DataFrame(componentDefinitionDetails, columns=["Component Definition Name", "Component Definition Ref", "Component Definition Desc", "Category Name", "Category Ref", "Risk Patterns"])

  dfmSupStandard=pd.DataFrame(supportedStandardDetails, columns=["Supported Standard Name", "Supported Standard Ref"])

  # We join the thre dataframes of library details in one dataframe
  dfmInfo = pd.concat([dfmLibrary, dfmComponentDefinitions, dfmSupStandard], axis=1, sort=False)

  rules = getRulesData(rootClass)
  dfmRules=pd.DataFrame(rules, columns=[
    "Rule Name", "Module", "Generated by GUI",
    "Condition Name", "Condition Value", "Condition Field",
    "Action Name", "Action Value", "Action Project"
    ])

  return dfm, dfmInfo, dfmRules


def getProductDataFromXml(xmlPath):
  table = list()
  # Main headers for the Excel
  table.append(
    ["Risk Pattern", "", "", "Use case", "", "" "Threat", "", "", "", "Weakness", "", "", "", "Countermeasure", "", "", "", "",
     ""])
  table.append(
    ["Id", "Name", "Desc", "Id", "Name", "Desc", "Id", "Name", "Desc", "References", "Id", "Name", "Desc", "References", "Id", "Name",
     "Desc", "Test steps", "References", "Standards"
     ])
  # We get all nodes of the xml file
  rootClass = supermod.parse(xmlPath, silence=True)
  # we get all important data from the xml file and transform it in a list
  table.extend(getAllDataFromRiskPatterns(rootClass))

  # These are the column names of the DataFrame
  columns = ["Risk Pattern Id", "Risk Pattern Name", "Risk Pattern Desc", "Use case Id", "Use case Name", "Use case Desc", "Threat Id", "Threat Name",
             "Threat Desc", "Threat Refs", "Weakness Id", "Weakness Name", "Weakness Desc", "Weakness Refs",
             "Countermeasure Id", "Countermeasure Name", "Countermeasure Desc", "Countermeasure Test steps",
             "Countermeasure Refs", "Countermeasure standards"]

  dfm = pd.DataFrame(table, columns=columns)

  tableInfo = list()
  tableInfo.append(["Trust Zone", "Component", "Question answered"])

  root = etree.parse(xmlPath).getroot()

  trustzones = dict()
  for trustzone in root.find("trustZones").iter("trustZone"):
    trustzones[trustzone.attrib['ref']] = trustzone.attrib['name']

  for component in root.iter("component"):
    for trustZone in component.iter("trustZone"):
      tz = trustZone.attrib["ref"]
    compName = component.attrib["name"]

    tz = trustzones[tz]
    comp = compName
    for question in component.iter("question"):
      tableInfo.append([tz, comp, question.attrib["ref"]])
      tz = ""
      comp = ""


  columnsQuestions = ["Trust Zone", "Component", "Question answered"]
  dfmQuestions = pd.DataFrame(tableInfo, columns=columnsQuestions)

  return dfm, dfmQuestions


def convertList(string, startTag, endTag, toTag):

  while string.find(startTag)!=-1:
    before=string[0:string.find(startTag)]
    string=string[string.find(startTag)+len(startTag):]
    transform=string[string.find("<li>"):string.find(endTag)]
    transform=transform.replace("<li> ", toTag)
    transform=transform.replace("<li>", toTag)
    transform=transform.replace("</li>","\n")
    after=string[string.find(endTag)+len(endTag):]
    string=before+transform+after
  return string

# With this function, we want to remove from the text, those tags are not necessary for the Excel
def convertFormatText(string):
  string=string.replace("&lt;","<")
  string=string.replace("&gt;",">")
  string=string.replace(" \n","")
  string=string.replace("\n","")
  string=convertList(string, "<ul>", "</ul>", "* ")
  string=convertList(string, "<ol>", "</ol>", "# ")
  string=string.replace("\t","")

  string=string.replace("&#8217;","'") 
  string=string.replace("&#8212;","--")
  string=string.replace("&#8220;","<")
  string=string.replace("&#8221;",">")
  string=string.replace("&#8230;","...")  
  string=string.replace("&quot;","'")
  string=string.replace("<div> ","")
  string=string.replace("<div>","")
  string=string.replace("</div>","\n")
  string=string.replace("<br />","")
  string=string.replace("<br/>","")
  string=string.replace("&nbsp;","")
  string=string.replace("<b>","**")
  string=string.replace("</b>","**")
  string=string.replace("<strong>","**")
  string=string.replace("</strong>","**")
  string=string.replace("<code>","@code@")
  string=string.replace("</code>","")
  string=string.replace("<div style=\"\">","")
  string=string.replace("<blockquote>","  ")
  string=string.replace("</blockquote>","")
  string=string.replace("<p> ","")
  string=string.replace("<p>","")
  string=string.replace("</span>","")
  string=string.replace('<span class="caps">',"%span%")
  string=string.replace("</p>","\n")
  string=string.replace("\n ","\n")
  string=string.replace("\n\n","\n")
  
  while(string.find("  ")!=-1):
    string=string.replace("  "," ")
  string=string.strip()

  return string
# We convert the references to one format [name|url]
def convertRefsToString(refs):
  text=""
  for ref in refs.get_reference():
    text+="[%s|%s]\n" %(ref.get_name(), ref.get_url())
  return text

# We convert the standards to one format [supportedStandardName|ref]
def convertStandardsToString(standards,supportedStandards):
  text=""
  for standard in standards.get_standard():
    supportedStandardRef=standard.get_supportedStandardRef()
    try:
      for supportedStandard in supportedStandards.get_supportedStandard():
        if supportedStandard.get_ref()== supportedStandardRef:
          supportedStandardRef=supportedStandard.get_ref()
          text+="[%s|%s]\n" %(supportedStandardRef, standard.get_ref())
    except:
      text += "[%s|%s]\n" % (supportedStandardRef, standard.get_ref())
  return text[0:-1]

# We apply a specific format to the sheet Library properties
def applyFormatLibraryProperties(excelPath):
  workbook = load_workbook(excelPath)
  sheetName="Library properties"
  ws = workbook.get_sheet_by_name(sheetName)
  ws.sheet_view.showGridLines = False
  
  columns=[
    [1, 2],             # Columns for the general info of the library
    [3, 4, 5, 6, 7,8],  # Columns for the Component Definitions
    [9, 10]             # Columns for the Supported Standards
  ]
  # Header colors for the library info, Component Definitions and Supported Standards
  headerColors=["2a6099", "ff9900", "800080"]
  colors=[
    ["dee6ef", "b4c7dc"], # Colors for the Library info
    ["ffdbb6", "ffb66c"], # Colors for the Component Definitions
    ["e0c2cd", "bf819e"]  # Colors for the Supported Standards
  ]
  for i in range(0,len(headerColors)):
    applyStylesTo(worksheet=ws, excelPath=excelPath, sheetName=sheetName, color=headerColors[i], columns=columns[i], headerRows=1)
    mergeCells(worksheet=ws, columns=columns[i], rowsHeader=1, numMaxRows=getMaxRowsInColumns(excelPath, sheetName), colors=colors[i], merge=False)
  
  workbook.save(excelPath)
# We use this method to apply a background color to the cells by columns. Now, we use it only for the headers.
def applyStylesTo(worksheet, excelPath, sheetName, color, columns, headerRows):
  for col in columns:
    worksheet.column_dimensions[get_column_letter(col)].width = 30    
    for row in range(1,headerRows+1): 
      setStylesHeader(cell=worksheet.cell(row=row, column=col), bgColor=color)

def applyFormatLibraryRules(excelPath):
  workbook = load_workbook(excelPath)
  sheetName="Rules"
  ws = workbook.get_sheet_by_name(sheetName)
  ws.sheet_view.showGridLines = False
  
  columns=[
    [1, 2, 3],                    # Columns for the rule details
    [4, 5, 6, 7, 8, 9],       # Columns for the condition details
    [10, 11, 12, 13, 14, 15]  # Columns for the action details
  ]
  # Header colors for the library info, Component Definitions and Supported Standards
  headerColors=["2a6099", "ff9900", "800080"]
  colors=[
    ["dee6ef", "b4c7dc"], # Colors for the rule details
    ["ffdbb6", "ffb66c"], # Colors for the condition details
    ["e0c2cd", "bf819e"]  # Colors for the action details
  ]
  merge=[True, False, False]
  max_rows=getMaxRowsInColumns(excelPath, sheetName)

  for i in range(0,len(headerColors)):
    applyStylesTo(worksheet=ws, excelPath=excelPath, sheetName=sheetName, color=headerColors[i], columns=columns[i], headerRows=1)
    mergeCells(worksheet=ws, columns=columns[i], rowsHeader=1, numMaxRows=max_rows, colors=colors[i], merge=merge[i])
  
  workbook.save(excelPath)


def applyFormatProductQuestions(excelPath):
  workbook = load_workbook(excelPath)
  sheetName = "Questions"
  ws = workbook.get_sheet_by_name(sheetName)
  ws.sheet_view.showGridLines = False

  columns = [
    [1],
    [2],
    [3]
  ]
  headerColors = ["2a6099", "ff9900", "800080"]
  colors = [
    ["dee6ef", "b4c7dc"],
    ["ffdbb6", "ffb66c"],
    ["e0c2cd", "bf819e"]
  ]
  merge = [True, True, True]
  max_rows = getMaxRowsInColumns(excelPath, sheetName)

  for i in range(0, len(headerColors)):
    applyStylesTo(worksheet=ws, excelPath=excelPath, sheetName=sheetName, color=headerColors[i], columns=columns[i],
                  headerRows=1)
    mergeCells(worksheet=ws, columns=columns[i], rowsHeader=1, numMaxRows=max_rows, colors=colors[i], merge=merge[i])

  workbook.save(excelPath)


# We apply format to the risk pattern sheet
def applyFormatLibrary(excelPath, sheetName):
  workbook = load_workbook(excelPath)
  ws = workbook.get_sheet_by_name(sheetName)
  ws.sheet_view.showGridLines = False
  # We get the number of rows for the sheet "Risk Patterns"
  max_rows=getMaxRowsInColumns(excelPath, sheetName)
  colors=[
    ["dee6ef", "b4c7dc"], # Colors for Risk Patterns
    ["e0c2cd", "bf819e"], # Colors for Use cases
    ["ffd7d7", "ffa6a6"], # Colors for Threats
    ["ffdbb6", "ffb66c"], # Colors for Weaknesses
    ["dde8cb", "afd095"]  # Colors for Countermeasures
  ]
  columnsList=[
    [1, 2, 3],                # Columns for Risk Patterns
    [4, 5, 6],                      # Columns for Use cases
    [7, 8, 9, 10],             # Columns for Threats
    [11, 12, 13, 14],          # Columns for Weaknesses
    [15, 16, 17, 18, 19, 20]  # Columns for Countermeasures
  ]
  # Header color for Risk Patterns, Use cases, Threats, Weaknesses and Countermeasures (in this order)
  colorsHeader=["2a6099", "800080", "ff0000", "ff9900", "38761d"]
  # Here we apply the format, first for the headers with the two first formats and then for the rest of the table
  for i in range(0, len(colors)):
    mergeHeaders(ws, [min(columnsList[i]), max(columnsList[i])])
    applyStylesTo(worksheet=ws, excelPath=excelPath, sheetName=sheetName, color=colorsHeader[i], columns=columnsList[i], headerRows=2)
    mergeCells(worksheet=ws, columns=columnsList[i], rowsHeader=2, numMaxRows=max_rows, colors=colors[i], merge=True)
   
  workbook.save(excelPath)

# In this method, we merge several cells (vertical) and apply the correspoding format.
def mergeCells(worksheet, columns, rowsHeader, numMaxRows, colors, merge):
  rowsToMerge=list()
  firstRow=""
  if numMaxRows > rowsHeader+1:
    # First we get the cells to merge
    for row in range(rowsHeader+1, numMaxRows):
      cell = worksheet.cell(row=row, column=columns[0]).value
      if firstRow == "":
        firstRow=row
      else:
        if cell != None and cell != "":
          rowsToMerge.append([firstRow, row-1])
          firstRow=row
    rowsToMerge.append([firstRow, numMaxRows-1])    
    # Then we merge the cells and we apply the format to this cells
    for col in columns:
      cont=0
      for rows in rowsToMerge:
        setStyles(cell=worksheet.cell(row=rows[0], column=col), bgColor=colors[cont%2]) 
        if merge and len(rowsToMerge)> 1:     
          worksheet.merge_cells(start_row=rows[0], start_column=col, end_row=rows[1], end_column=col)
        cont+=1
      if len(rowsToMerge) <= 1 and worksheet.title == "Risk Patterns":  
        increase=0
        for row in range(rowsHeader+1, numMaxRows):
          setStyles(cell=worksheet.cell(row=row, column=col), bgColor=colors[increase%2]) 
          increase+=1
        worksheet.column_dimensions[get_column_letter(col)].hidden= True

# In this method we merge the cells (horizontal) for the headers
def mergeHeaders(worksheet, columnsToMerge):
  worksheet.merge_cells(start_row=1, start_column=columnsToMerge[0], end_row=1, end_column=columnsToMerge[1])

# We get the number of rows for one column name
def getMaxRowsInColumns(excelPath, sheetName):
  dfm = pd.read_excel(excelPath, sheetName)
  return len(dfm.index)+2

# Here we apply a default styles to the cell and apply a certain background color
def setStyles(cell, bgColor): 
  cell.font = Font(bold=False, color='000000')
  bd = Side(style='thin', color=Color('FFFFFF'))
  cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
  cell.fill = PatternFill(fill_type='solid', fgColor=Color(bgColor))
  cell.alignment = Alignment(shrinkToFit=True, horizontal='left', wrapText=True, vertical='center')

 
# Here we apply a default styles to the header cell and apply a certain background color
def setStylesHeader(cell, bgColor):
  cell.font = Font(bold=True, color='FFFFFF')
  bd = Side(style='thin', color='FFFFFF')
  cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
  cell.fill = PatternFill(fill_type='solid', fgColor=Color(bgColor))
  cell.alignment = Alignment(shrinkToFit=True, horizontal='center', wrapText=True, vertical='center')

def convertXmlToExcel(xmlPath, excelPath):
  #We read the data from the xml and get the necessary DataFrames to build the Excel file.
  (data,info, rules)=getDataFromXml(str(xmlPath))
  # We create the new Excel file and we open it to write on it.
  _log.info("Excel file opened to be written: %s." %excelPath)
  writer = pd.ExcelWriter(str(excelPath))
  # We write the data for the risk patterns sheet. 
  sheetName='Risk Patterns'
  data.to_excel(excel_writer=writer,sheet_name=sheetName, header=None, index=False, merge_cells=True)
  _log.info("Data from xml put into the sheet: %s." %sheetName)
  # We write the data for the Rules sheet. 
  rules.to_excel(excel_writer=writer,sheet_name='Rules', index=False)
  _log.info("Library info from xml put into the sheet 'Rules'.")

  # We write the data for the Library properties sheet. 
  info.to_excel(excel_writer=writer,sheet_name='Library properties', index=False)
  _log.info("Library info from xml put into the sheet 'Library properties'.")
  
  # We save the changes for the Excel
  writer.save()

  # We apply format to both sheets
  applyFormatLibraryProperties(excelPath)
  _log.info("Format applied to the sheet 'Library properties'.")
  applyFormatLibrary(excelPath, sheetName)
  _log.info("Format applied to the sheet '%s'." %sheetName)
  applyFormatLibraryRules(excelPath)
  _log.info("Format applied to the sheet 'Rules'.")
  print("Excel file generated in the path: '%s'." %excelPath)
  return excelPath


def convertProductXmlToExcel(xmlPath, excelPath):
  # We read the data from the xml and get the necessary DataFrames to build the Excel file.
  data, dataQ = getProductDataFromXml(str(xmlPath))
  # We create the new Excel file and we open it to write on it.
  _log.info("Excel file opened to be written: %s." % excelPath)
  writer = pd.ExcelWriter(str(excelPath))
  # We write the data for the risk patterns sheet.
  sheetName = 'Risk Patterns'
  data.to_excel(excel_writer=writer, sheet_name=sheetName, header=None, index=False, merge_cells=True)
  _log.info("Data from xml put into the sheet: %s." % sheetName)

  dataQ.to_excel(excel_writer=writer, sheet_name="Questions", header=None, index=False, merge_cells=True)

  writer.save()

  # We apply format to both sheets
  # applyFormatLibraryProperties(excelPath)
  # _log.info("Format applied to the sheet 'Library properties'.")
  applyFormatLibrary(excelPath, sheetName)
  _log.info("Format applied to the sheet '%s'." % sheetName)
  applyFormatProductQuestions(excelPath)
  _log.info("Format applied to the sheet 'Questions'.")
  print("Excel file generated in the path: '%s'." % excelPath)
  return excelPath


def main():
  # We load the different libraries from the folder 'libraries'
  libraries=os.listdir(str(Path.cwd() / "libraries"))
  # We ask the user, which xml from the libraries directory want to transform
  text="Select the library (number of the library) to get the Excel file:\n"
  for library in libraries:
    if library.endswith(".xml"):
      text+="%i - %s\n" %(libraries.index(library),library)
  value=-1
  while(value < 0 or value > len(libraries)):
    value=int(input(text))
  _log.info("Library selected: %s." %libraries[value])

  xmlPath=str(Path.cwd() / "libraries" / libraries[value])
  excelPath=str(Path.cwd() / "outFiles" / "spreadSheets" / libraries[value].replace(".xml",".xlsx"))
  # We launch the method to convert the xml file to Excel
  convertXmlToExcel(xmlPath, excelPath)
  

if __name__ == '__main__':
  main()